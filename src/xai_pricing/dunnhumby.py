import hashlib
import shutil
import urllib.request
import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .config import (
    RAW_DATA_DIR,
    SOURCE_ARCHIVE_NAME,
    SOURCE_ARCHIVE_SHA256,
    SOURCE_GUIDE_NAME,
    SOURCE_URL,
    SOURCE_WORKBOOK_NAME,
)


@dataclass(frozen=True)
class SourcePaths:
    archive_path: Path
    workbook_path: Path
    guide_path: Path


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def ensure_source_download() -> SourcePaths:
    RAW_DATA_DIR.mkdir(parents=True, exist_ok=True)
    archive_path = RAW_DATA_DIR / SOURCE_ARCHIVE_NAME

    if archive_path.exists() and sha256_file(archive_path) != SOURCE_ARCHIVE_SHA256:
        archive_path.unlink()

    if not archive_path.exists():
        tmp_path = archive_path.with_suffix(".tmp")
        with urllib.request.urlopen(SOURCE_URL) as response, tmp_path.open("wb") as output:
            shutil.copyfileobj(response, output)
        actual_sha = sha256_file(tmp_path)
        if actual_sha != SOURCE_ARCHIVE_SHA256:
            tmp_path.unlink(missing_ok=True)
            raise RuntimeError(
                "Downloaded archive hash mismatch. "
                f"Expected {SOURCE_ARCHIVE_SHA256}, got {actual_sha}."
            )
        tmp_path.replace(archive_path)

    extracted_dir = RAW_DATA_DIR / archive_path.stem
    workbook_path = extracted_dir / SOURCE_WORKBOOK_NAME
    guide_path = extracted_dir / SOURCE_GUIDE_NAME

    if not workbook_path.exists() or not guide_path.exists():
        with zipfile.ZipFile(archive_path) as archive:
            archive.extractall(RAW_DATA_DIR)
        workbook_matches = list(RAW_DATA_DIR.rglob(SOURCE_WORKBOOK_NAME))
        guide_matches = list(RAW_DATA_DIR.rglob(SOURCE_GUIDE_NAME))
        if workbook_matches:
            workbook_path = workbook_matches[0]
        if guide_matches:
            guide_path = guide_matches[0]

    return SourcePaths(
        archive_path=archive_path,
        workbook_path=workbook_path,
        guide_path=guide_path,
    )


def open_workbook(path: Path):
    return load_workbook(path, read_only=True, data_only=True)


def sheet_headers(sheet, header_row: int = 2) -> list[str]:
    headers = []
    for value in next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)):
        text = str(value).strip() if value is not None else ""
        headers.append(text)
    return headers
